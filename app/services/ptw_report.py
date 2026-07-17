"""PTW close-out report (fpdf2) + permit register export (openpyxl).

The close-out report is the "golden thread" document the closed-loop PTW
system exists to produce: the permit certificate, the full evidence
timeline (every lifecycle action with actor, timestamp, GPS fix, drawn
signature and onsite photos), the approvals register, gas-test log,
isolation log, suspension/extension history, the Work Completed
declaration + outcome, the handback inspection checklist, and the
closure sign-off — pulled straight from the same rows the tamper-evident
audit hash-chain covers. A permit that is not yet CLOSED renders with a
PROVISIONAL watermark on every page.

Pure-Python (fpdf2 core fonts, latin-1) — same conventions as
app/services/report_pdf.py.
"""

from __future__ import annotations

import base64
import io
from datetime import datetime
from typing import Any

from fpdf import FPDF

_REPL = {"—": "-", "–": "-", "‘": "'", "’": "'", "“": '"', "”": '"',
         "•": "*", "₹": "Rs ", "→": "->", "≥": ">=", "≤": "<=", " ": " "}


def _s(text: Any) -> str:
    """Sanitise to latin-1 (fpdf2 core fonts) — map common Unicode then drop the rest."""
    t = str(text if text is not None else "—")
    for k, v in _REPL.items():
        t = t.replace(k, v)
    return t.encode("latin-1", "replace").decode("latin-1")


NAVY = (30, 41, 90)
GREY = (100, 100, 100)
LIGHT = (235, 235, 240)
RED = (192, 57, 43)
GREEN = (39, 139, 87)
AMBER = (230, 126, 34)

_ACTION_LABELS = {
    "APPROVE_ISSUER": "Issuer Approval",
    "APPROVE_SAFETY": "Safety Officer Approval",
    "APPROVE_PLANT_HEAD": "Plant Head Approval",
    "APPROVE": "Approval",
    "ISSUE": "Permit Issued",
    "ACCEPT": "Receiver Acceptance",
    "ISOLATION_VERIFY": "Isolation Verified",
    "SUSPEND": "Suspended",
    "RESUME": "Resumed",
    "EXTEND": "Extension",
    "WORK_COMPLETED_DECLARE": "Work Completed Declared",
    "HANDBACK_INSPECT": "Handback Inspection",
    "CLOSE": "Closure Approval",
    "CANCEL": "Cancelled",
    "REJECT": "Rejected",
}


def _fmt_dt(dt: datetime | None) -> str:
    if dt is None:
        return "—"
    return dt.strftime("%d %b %Y %H:%M UTC")


def _fmt_gps(lat: float | None, lng: float | None, acc: float | None) -> str:
    if lat is None or lng is None:
        return "no GPS fix recorded"
    base = f"{lat:.6f}, {lng:.6f}"
    if acc is not None:
        base += f" (±{acc:.0f} m)"
    return base


class _PtwReport(FPDF):
    def __init__(self, permit_number: str, provisional: bool):
        super().__init__(orientation="P", unit="mm", format="A4")
        self.permit_number = permit_number
        self.provisional = provisional
        self.set_auto_page_break(auto=True, margin=22)
        self.set_title(_s(f"Permit Close-out Report {permit_number}"))

    def cell(self, *a, **k):  # type: ignore[override]
        if len(a) >= 3 and isinstance(a[2], str):
            a = (a[0], a[1], _s(a[2])) + a[3:]
        for key in ("txt", "text"):
            if key in k and isinstance(k[key], str):
                k[key] = _s(k[key])
        return super().cell(*a, **k)

    def multi_cell(self, *a, **k):  # type: ignore[override]
        if len(a) >= 3 and isinstance(a[2], str):
            a = (a[0], a[1], _s(a[2])) + a[3:]
        for key in ("txt", "text"):
            if key in k and isinstance(k[key], str):
                k[key] = _s(k[key])
        return super().multi_cell(*a, **k)

    def header(self):  # noqa: D102
        self.set_font("Helvetica", "B", 9)
        self.set_text_color(*NAVY)
        self.cell(0, 6, f"SafeOps360 — Permit to Work Close-out Report — {self.permit_number}", align="L")
        self.ln(7)
        self.set_draw_color(*NAVY)
        self.set_line_width(0.4)
        self.line(10, 17, 200, 17)
        if self.provisional:
            # PROVISIONAL watermark on every page until the permit is CLOSED.
            self.set_font("Helvetica", "B", 52)
            self.set_text_color(245, 200, 200)
            with self.rotation(45, 105, 150):
                self.text(45, 160, "PROVISIONAL")
            self.set_text_color(0, 0, 0)
        self.set_y(21)

    def footer(self):  # noqa: D102
        self.set_y(-16)
        self.set_font("Helvetica", "I", 7.5)
        self.set_text_color(*GREY)
        self.cell(
            0, 5,
            "Confidential — generated from tamper-evident audit records (SHA-256 hash chain).",
            align="L",
        )
        self.cell(0, 5, f"Page {self.page_no()}/{{nb}}", align="R")

    # ── layout helpers ──
    def section(self, title: str):
        if self.get_y() > 255:
            self.add_page()
        self.ln(3)
        self.set_font("Helvetica", "B", 11)
        self.set_text_color(*NAVY)
        self.cell(0, 7, title)
        self.ln(8)
        self.set_text_color(0, 0, 0)

    def kv(self, label: str, value: Any, w_label: float = 48):
        self.set_font("Helvetica", "B", 8.5)
        self.set_text_color(*GREY)
        self.cell(w_label, 5.5, label)
        self.set_font("Helvetica", "", 9)
        self.set_text_color(0, 0, 0)
        self.multi_cell(0, 5.5, str(value if value not in (None, "") else "—"))

    def table_header(self, cols: list[tuple[str, float]]):
        self.set_font("Helvetica", "B", 8)
        self.set_fill_color(*LIGHT)
        self.set_text_color(*NAVY)
        for name, w in cols:
            self.cell(w, 6, name, border=1, fill=True)
        self.ln(6)
        self.set_text_color(0, 0, 0)

    def table_row(self, cells: list[tuple[str, float]], h: float = 5.5):
        self.set_font("Helvetica", "", 8)
        if self.get_y() > 265:
            self.add_page()
        for text, w in cells:
            self.cell(w, h, (text or "—")[:60], border=1)
        self.ln(h)


def render_ptw_closeout_pdf(data: dict[str, Any]) -> bytes:
    """`data` is the fully-resolved dict built by build_ptw_report_data()."""
    p = data["permit"]
    provisional = p["status"] != "CLOSED"
    pdf = _PtwReport(p["number"], provisional)
    pdf.alias_nb_pages()
    pdf.add_page()

    # ── Cover block ────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 17)
    pdf.set_text_color(*NAVY)
    pdf.cell(0, 10, f"Permit Close-out Report — {p['number']}")
    pdf.ln(11)
    status_color = GREEN if p["status"] == "CLOSED" else (RED if p["status"] in ("CANCELLED", "REJECTED") else AMBER)
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(*status_color)
    pdf.cell(0, 6, f"Status: {p['status']}" + (f"   |   Outcome: {p['outcome']}" if p.get("outcome") else ""))
    pdf.ln(9)
    pdf.set_text_color(0, 0, 0)

    pdf.section("1. Permit Certificate")
    pdf.kv("Permit No.", p["number"])
    pdf.kv("Type", p["type"])
    pdf.kv("Plant / Location", f"{p['plantName']} — {p['location']}" + (f" ({p['specificLocation']})" if p.get("specificLocation") else ""))
    pdf.kv("Scope of Work", p["scopeOfWork"])
    pdf.kv("Validity", f"{_fmt_dt(p['validFrom'])}  ->  {_fmt_dt(p['validTo'])}")
    pdf.kv("Originator", p["originatorName"])
    pdf.kv("Issuer", p["issuerName"])
    pdf.kv("Receiver", p["receiverName"])
    if p.get("contractorName"):
        pdf.kv("Contractor", p["contractorName"])
    pdf.kv("FLRA required", "Yes" if p["flraRequired"] else "No (waived by instance policy)")
    if p.get("gpsLatitude") is not None:
        pdf.kv("Creation GPS", _fmt_gps(p["gpsLatitude"], p["gpsLongitude"], None))

    # ── Crew ───────────────────────────────────────────────────────────
    if data["crew"]:
        pdf.section("2. Work Crew")
        pdf.table_header([("Name", 60), ("Role", 35), ("Training OK", 25), ("PPE OK", 25), ("Removed", 45)])
        for c in data["crew"]:
            pdf.table_row([
                (c["name"], 60), (c["role"], 35),
                ("Yes" if c["trainingValid"] else ("No" if c["trainingValid"] is False else "—"), 25),
                ("Yes" if c["ppeValid"] else ("No" if c["ppeValid"] is False else "—"), 25),
                (_fmt_dt(c["removedAt"]) if c["removedAt"] else "—", 45),
            ])

    # ── Evidence timeline — the closed loop ────────────────────────────
    pdf.section("3. Lifecycle Evidence Timeline (GPS + photo + signature)")
    if not data["evidence"]:
        pdf.set_font("Helvetica", "I", 9)
        pdf.cell(0, 6, "No evidence rows recorded (permit predates the closed-loop rebuild).")
        pdf.ln(7)
    for ev in data["evidence"]:
        if pdf.get_y() > 235:
            pdf.add_page()
        pdf.set_font("Helvetica", "B", 9.5)
        pdf.set_text_color(*NAVY)
        pdf.cell(120, 6, _ACTION_LABELS.get(ev["action"], ev["action"]))
        pdf.set_font("Helvetica", "", 8.5)
        pdf.set_text_color(*GREY)
        pdf.cell(0, 6, _fmt_dt(ev["capturedAt"]), align="R")
        pdf.ln(6)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "", 8.5)
        pdf.multi_cell(0, 4.8, f"By {ev['actorName']}   |   GPS: {_fmt_gps(ev['gpsLatitude'], ev['gpsLongitude'], ev['gpsAccuracyMeters'])}")
        if ev.get("declarationText"):
            pdf.set_font("Helvetica", "I", 8.5)
            pdf.multi_cell(0, 4.8, f'Declaration: "{ev["declarationText"]}"')
        if ev.get("comments"):
            pdf.set_font("Helvetica", "", 8.5)
            pdf.multi_cell(0, 4.8, f"Notes: {ev['comments']}")
        # Drawn signature (data-URL PNG) inline.
        sig = ev.get("signatureImageBase64")
        if sig:
            try:
                raw = base64.b64decode(sig.split(",", 1)[1] if "," in sig else sig)
                y = pdf.get_y()
                if y > 250:
                    pdf.add_page()
                    y = pdf.get_y()
                pdf.image(io.BytesIO(raw), x=14, y=y + 1, h=12)
                pdf.set_y(y + 14)
                pdf.set_font("Helvetica", "I", 7)
                pdf.set_text_color(*GREY)
                pdf.cell(0, 4, f"Signed by {ev['actorName']}")
                pdf.ln(5)
                pdf.set_text_color(0, 0, 0)
            except Exception:  # noqa: BLE001 — a bad signature blob never kills the report
                pdf.set_font("Helvetica", "I", 8)
                pdf.cell(0, 5, "[signature on file - could not render]")
                pdf.ln(5)
        # Onsite photos.
        photos = ev.get("photoBytes") or []
        if photos:
            x = 14
            y = pdf.get_y() + 1
            if y > 230:
                pdf.add_page()
                y = pdf.get_y() + 1
            rendered = 0
            for blob in photos[:3]:
                try:
                    pdf.image(io.BytesIO(blob), x=x, y=y, h=32)
                    x += 62
                    rendered += 1
                except Exception:  # noqa: BLE001
                    continue
            if rendered:
                pdf.set_y(y + 34)
            extra = len(photos) - rendered
            if extra > 0:
                pdf.set_font("Helvetica", "I", 7.5)
                pdf.set_text_color(*GREY)
                pdf.cell(0, 4, f"+{extra} more photo(s) on file")
                pdf.ln(5)
                pdf.set_text_color(0, 0, 0)
        pdf.ln(2)

    # ── Approvals register ─────────────────────────────────────────────
    if data["approvals"]:
        pdf.section("4. Approvals Register")
        pdf.table_header([("Step", 45), ("Approver", 55), ("Decision", 25), ("When", 40), ("Comments", 25)])
        for a in data["approvals"]:
            pdf.table_row([
                (a["step"], 45), (a["approverName"], 55), (a["decision"], 25),
                (_fmt_dt(a["decidedAt"]), 40), ((a["comments"] or "—")[:24], 25),
            ])

    # ── Gas tests ──────────────────────────────────────────────────────
    if data["gasReadings"]:
        pdf.section("5. Gas Test Log")
        pdf.table_header([("When", 38), ("By", 42), ("Readings", 70), ("Exceedance", 22), ("Pre-entry", 18)])
        for g in data["gasReadings"]:
            readings = ", ".join(
                f"{r.get('parameter')}={r.get('value')}" for r in (g["readings"] or [])
            )
            pdf.table_row([
                (_fmt_dt(g["recordedAt"]), 38), (g["byName"], 42), (readings, 70),
                ("YES" if g["isExceedance"] else "no", 22), ("yes" if g["isPreEntry"] else "no", 18),
            ])

    # ── Isolations ─────────────────────────────────────────────────────
    if data["isolations"]:
        pdf.section("6. Isolation / LOTO Log")
        pdf.table_header([("Point / Tag", 45), ("Type", 30), ("Verified", 45), ("Restored", 45), ("LOTO", 25)])
        for i in data["isolations"]:
            pdf.table_row([
                (i["tag"], 45), (i["type"], 30),
                (f"{_fmt_dt(i['verifiedAt'])} {i['verifiedByName'] or ''}".strip(), 45),
                (f"{_fmt_dt(i['restoredAt'])} {i['restoredByName'] or ''}".strip(), 45),
                (i["lotoTag"] or "—", 25),
            ])

    # ── Suspensions & extensions ───────────────────────────────────────
    if data["suspensions"]:
        pdf.section("7. Suspension / Resumption History")
        pdf.table_header([("Suspended", 40), ("Reason", 55), ("Resumed", 40), ("Re-FLRA", 20), ("By", 35)])
        for s in data["suspensions"]:
            pdf.table_row([
                (_fmt_dt(s["suspendedAt"]), 40), ((s["reasonDetail"] or s["reason"])[:52], 55),
                (_fmt_dt(s["resumedAt"]), 40), ("yes" if s["reFlraRequired"] else "no", 20),
                (s["byName"], 35),
            ])
    if data["extensions"]:
        pdf.section("8. Validity Extensions")
        pdf.table_header([("Requested", 40), ("New Valid To", 40), ("Status", 25), ("Decided by", 45), ("Reason", 40)])
        for e in data["extensions"]:
            pdf.table_row([
                (_fmt_dt(e["requestedAt"]), 40), (_fmt_dt(e["newValidTo"]), 40),
                (e["status"], 25), (e["approverName"] or "—", 45), ((e["reason"] or "")[:38], 40),
            ])

    # ── Work Completed + Handback + Closure ────────────────────────────
    pdf.section("9. Work Completed Declaration")
    pdf.kv("Declared at", _fmt_dt(p["workCompletedAt"]))
    pdf.kv("Declared by", p["workCompletedByName"])
    pdf.kv("Outcome", p.get("outcome") or "—")
    pdf.kv("Notes", p.get("returnNotes") or "—")

    pdf.section("10. Handback Inspection")
    pdf.kv("Inspected at", _fmt_dt(p["siteVerifiedAt"]))
    pdf.kv("Inspected by", p["siteVerifiedByName"])
    for item, ok in (p.get("siteVerificationChecklist") or {}).items():
        pdf.kv(f"  - {item}", "PASS" if ok else "FAIL")

    pdf.section("11. Closure")
    pdf.kv("Closed at", _fmt_dt(p["closedAt"]))
    pdf.kv("Closed by", p["closedByName"])
    pdf.kv("Closing remark", p.get("closingRemark") or "—")
    if p.get("cancelledAt"):
        pdf.kv("Cancelled at", _fmt_dt(p["cancelledAt"]))
        pdf.kv("Cancelled by", p["cancelledByName"])
        pdf.kv("Cancellation reason", p.get("cancellationReason"))

    # ── Integrity stamp ────────────────────────────────────────────────
    pdf.section("12. Record Integrity")
    pdf.set_font("Helvetica", "", 8.5)
    pdf.multi_cell(
        0, 5,
        "Every row rendered above is captured in SafeOps360's tamper-evident "
        "audit trail (per-entity SHA-256 hash chain, weekly integrity job). "
        f"Latest chain entry for this permit: {data.get('latestAuditHash') or 'n/a'}",
    )

    out = pdf.output()
    return bytes(out)


def build_register_xlsx(rows: list[dict[str, Any]]) -> bytes:
    """Permit register export — one row per permit."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Permit Register"

    headers = [
        "Permit No.", "Type", "Status", "Outcome", "Plant", "Location",
        "Scope of Work", "Valid From", "Valid To", "Originator", "Issuer",
        "Receiver", "Contractor", "FLRA Required", "Issued At", "Accepted At",
        "Work Completed At", "Handback At", "Closed At", "Closing Remark",
        "Suspensions", "Extensions", "Gas Exceedances", "Archived",
    ]
    header_fill = PatternFill("solid", fgColor="1E295A")
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill

    def _d(v):
        return v.strftime("%Y-%m-%d %H:%M") if isinstance(v, datetime) else (v or "")

    for r_i, r in enumerate(rows, start=2):
        values = [
            r["number"], r["type"], r["status"], r.get("outcome") or "",
            r["plantName"], r["location"], r["scopeOfWork"],
            _d(r["validFrom"]), _d(r["validTo"]),
            r["originatorName"], r["issuerName"], r["receiverName"],
            r.get("contractorName") or "", "Yes" if r["flraRequired"] else "No",
            _d(r.get("issuedAt")), _d(r.get("activatedAt")),
            _d(r.get("workCompletedAt")), _d(r.get("siteVerifiedAt")),
            _d(r.get("closedAt")), r.get("closingRemark") or "",
            r.get("suspensionCount", 0), r.get("extensionCount", 0),
            r.get("exceedanceCount", 0), "Yes" if r.get("isArchived") else "No",
        ]
        for c_i, v in enumerate(values, start=1):
            ws.cell(row=r_i, column=c_i, value=v)

    widths = [16, 16, 14, 14, 18, 22, 40, 17, 17, 20, 20, 20, 18, 8, 17, 17, 17, 17, 17, 32, 6, 6, 6, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
